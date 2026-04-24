"""Stage 1: source-driven harvest with wave support and raw manifests."""

from __future__ import annotations

import asyncio
import json
import re
import xml.etree.ElementTree as ET
from datetime import UTC, datetime
from functools import lru_cache
from pathlib import Path
from typing import TYPE_CHECKING, Any

import aiohttp

from polisyos.batch_common.manifest import write_raw_manifest, write_stage_manifest
from polisyos.common.logger import get_logger
from polisyos.datasets.batch.checkpoints import hash_payload, load_json, write_json
from polisyos.datasets.batch.ckan_curation import curate_ckan_package
from polisyos.datasets.batch.normalizer import map_to_polisyos_metrics
from polisyos.datasets.metrics_map import load_metrics_map

if TYPE_CHECKING:
    from polisyos.datasets.batch.config import DatasetBatchConfig
    from polisyos.datasets.batch.source_registry import SourceSpec

logger = get_logger(__name__)
_HARVEST_MAX_PARALLELISM = 6
_SERIAL_HARVEST_SOURCES: frozenset[str] = frozenset(
    {
        "data_gov_ua_broad",
        "data_gov_ro_broad",
        "data_gov_md_broad",
        "data_gov_pl_broad",
    }
)
_SMOKE_PRIORITY_METRICS: tuple[str, ...] = (
    "gdp_per_capita",
    "gdp",
    "unemployment_rate",
    "inflation",
    "poverty_rate",
    "migration",
    "health_outcomes",
    "education_outcomes",
    "social_trust",
    "labor_force_participation",
    "institutional_quality",
)
_SOURCE_PRIORITY_METRICS: dict[str, tuple[str, ...]] = {
    "worldbank": (
        "gdp_per_capita",
        "poverty_rate",
        "health_outcomes",
        "institutional_quality",
        "education_outcomes",
        "labor_force_participation",
        "migration",
    ),
    "ilo": (
        "labor_force_participation",
        "unemployment_rate",
        "migration",
        "inflation",
        "gdp_per_capita",
        "education_outcomes",
    ),
    "who": (
        "health_outcomes",
        "life_expectancy",
        "poverty_rate",
    ),
    "wvs": ("social_trust",),
    "data_gov_ro_broad": (
        "municipal_budget",
        "education_outcomes",
        "health_outcomes",
        "unemployment_rate",
        "migration",
        "demography",
    ),
    "data_gov_ro": (
        "municipal_budget",
        "education_outcomes",
        "health_outcomes",
        "unemployment_rate",
        "migration",
        "demography",
    ),
    "data_gov_ro_exec": (
        "municipal_budget",
        "education_outcomes",
        "health_outcomes",
        "unemployment_rate",
        "migration",
        "demography",
    ),
    "data_gov_md_broad": (
        "municipal_budget",
        "education_outcomes",
        "health_outcomes",
        "unemployment_rate",
        "migration",
        "demography",
    ),
    "data_gov_md": (
        "municipal_budget",
        "education_outcomes",
        "health_outcomes",
        "unemployment_rate",
        "migration",
        "demography",
    ),
    "data_gov_md_exec": (
        "municipal_budget",
        "education_outcomes",
        "health_outcomes",
        "unemployment_rate",
        "migration",
        "demography",
    ),
    "data_gov_pl_broad": (
        "municipal_budget",
        "education_outcomes",
        "health_outcomes",
        "unemployment_rate",
        "migration",
        "demography",
    ),
    "data_gov_pl": (
        "municipal_budget",
        "education_outcomes",
        "health_outcomes",
        "unemployment_rate",
        "migration",
        "demography",
    ),
    "data_gov_pl_exec": (
        "municipal_budget",
        "education_outcomes",
        "health_outcomes",
        "unemployment_rate",
        "migration",
        "demography",
    ),
    "paris_opendata_exec": (
        "health_outcomes",
        "education_outcomes",
        "unemployment_rate",
        "migration",
    ),
    "nyc_opendata": ("health_outcomes", "education_outcomes", "migration"),
    "nyc_opendata_exec": ("health_outcomes", "education_outcomes", "migration"),
    "chicago_opendata": ("health_outcomes", "education_outcomes", "migration"),
    "chicago_opendata_exec": ("health_outcomes", "education_outcomes", "migration"),
    "opendatasoft_public": ("health_outcomes", "education_outcomes", "migration"),
}
_PRIORITY_PATTERNS: dict[str, tuple[str, ...]] = {
    "gdp_per_capita": ("gdp per capita", "gross domestic product per capita", "ввп на душу"),
    "gdp": ("gross domestic product", " gdp ", "ввп"),
    "unemployment_rate": (
        "unemployment",
        "jobless",
        "безробіт",
        "somaj",
        "șomaj",
        "bezroboc",
        "rynek pracy",
    ),
    "inflation": (
        "inflation",
        "consumer price",
        "cpi",
        "інфляц",
        "inflatie",
        "inflație",
        "inflacja",
    ),
    "poverty_rate": ("poverty", "deprivation", "бідн"),
    "migration": (
        "migration",
        "migrant",
        "refugee",
        "міграц",
        "migratie",
        "migrație",
        "migrac",
        "demografi",
    ),
    "health_outcomes": (
        "life expectancy",
        "healthy life expectancy",
        "mortality",
        "тривалість життя",
        "здоров",
        "sanat",
        "sănătate",
        "spital",
        "zdrow",
        "szpital",
    ),
    "education_outcomes": (
        "education",
        "enrollment",
        "enrolment",
        "school",
        "literacy",
        "освіт",
        "зарахув",
        "educat",
        "scoala",
        "școal",
        "elev",
        "edukac",
        "szkol",
        "uczni",
    ),
    "social_trust": ("social trust", "values survey", "довір"),
    "labor_force_participation": (
        "labor force participation",
        "labour force participation",
        "робочій силі",
    ),
    "institutional_quality": (
        "rule of law",
        "government effectiveness",
        "regulatory quality",
        "institutional quality",
    ),
    "municipal_budget": (
        "budget",
        "municipal budget",
        "local budget",
        "buget",
        "buget local",
        "venituri",
        "cheltuieli",
        "budzet",
        "budżet",
        "dochody",
        "wydatki",
    ),
    "demography": (
        "demography",
        "population",
        "birth",
        "death",
        "demograf",
        "populatie",
        "populație",
        "nasteri",
        "nașteri",
        "decese",
        "ludnosc",
        "ludność",
        "urodzenia",
        "zgony",
    ),
}
_PRIORITY_BLACKLIST: tuple[str, ...] = (
    "climate",
    "wildfire",
    "flood",
    "drought",
    "academic school year",
    "start month",
    "end month",
)
_WVS_STATIC_INDICATORS: tuple[dict[str, Any], ...] = (
    {
        "id": "A009",
        "name": "State of health (subjective)",
        "description": "WVS longitudinal subjective health item",
        "wave": "timeseries",
        "harvest_metric_candidates": ["health_outcomes"],
    },
    {
        "id": "A165",
        "name": "Most people can be trusted",
        "description": "WVS longitudinal social trust question",
        "wave": "timeseries",
        "harvest_metric_candidates": ["social_trust"],
    },
    {
        "id": "A170",
        "name": "Satisfaction with your life",
        "description": "WVS longitudinal life satisfaction item",
        "wave": "timeseries",
        "harvest_metric_candidates": ["social_capital"],
    },
    {
        "id": "A173",
        "name": "How much freedom of choice and control",
        "description": "WVS longitudinal freedom-of-choice item",
        "wave": "timeseries",
        "harvest_metric_candidates": ["cultural_cluster"],
    },
    {
        "id": "A214",
        "name": "I see myself as someone who is generally trusting",
        "description": "WVS longitudinal generalized trust item",
        "wave": "timeseries",
        "harvest_metric_candidates": ["social_trust"],
    },
    {
        "id": "D059",
        "name": "Men make better political leaders than women do",
        "description": "WVS longitudinal gender norms item",
        "wave": "timeseries",
        "harvest_metric_candidates": ["gender_equality"],
    },
    {
        "id": "E023",
        "name": "Interest in politics",
        "description": "WVS longitudinal political interest item",
        "wave": "timeseries",
        "harvest_metric_candidates": ["social_capital"],
    },
    {
        "id": "E025",
        "name": "Political action: Signing a petition",
        "description": "WVS longitudinal petition participation item",
        "wave": "timeseries",
        "harvest_metric_candidates": ["social_capital"],
    },
    {
        "id": "E035",
        "name": "Income equality",
        "description": "WVS longitudinal income equality preference item",
        "wave": "timeseries",
        "harvest_metric_candidates": ["inequality"],
    },
    {
        "id": "E069_07",
        "name": "Confidence: Parliament",
        "description": "WVS longitudinal institutional confidence item for parliament",
        "wave": "timeseries",
        "harvest_metric_candidates": ["public_trust", "institutional_quality"],
    },
    {
        "id": "E069_11",
        "name": "Confidence: The Government",
        "description": "WVS longitudinal institutional confidence item for government",
        "wave": "timeseries",
        "harvest_metric_candidates": ["public_trust", "institutional_quality"],
    },
    {
        "id": "E069_17",
        "name": "Confidence: Justice System/Courts",
        "description": "WVS longitudinal institutional confidence item for courts",
        "wave": "timeseries",
        "harvest_metric_candidates": ["judicial_quality", "institutional_quality"],
    },
    {
        "id": "E110",
        "name": "Satisfaction with the way democracy develops",
        "description": "WVS longitudinal democracy satisfaction item",
        "wave": "timeseries",
        "harvest_metric_candidates": ["democracy_quality"],
    },
    {
        "id": "E117",
        "name": "Political system: Having a democratic political system",
        "description": "WVS longitudinal democracy preference item",
        "wave": "timeseries",
        "harvest_metric_candidates": ["democracy_quality"],
    },
    {
        "id": "E196",
        "name": "Extent of political corruption",
        "description": "WVS longitudinal perceived political corruption item",
        "wave": "timeseries",
        "harvest_metric_candidates": ["corruption_level"],
    },
    {
        "id": "E233",
        "name": "Democracy: Women have the same rights as men",
        "description": "WVS longitudinal gender equality in democracy item",
        "wave": "timeseries",
        "harvest_metric_candidates": ["gender_equality"],
    },
    {
        "id": "E286",
        "name": "Social activism: Donating to a group or campaign",
        "description": "WVS longitudinal social activism donation item",
        "wave": "timeseries",
        "harvest_metric_candidates": ["social_capital"],
    },
    {
        "id": "F108",
        "name": "Government protects freedom",
        "description": "WVS longitudinal freedom protection item",
        "wave": "timeseries",
        "harvest_metric_candidates": ["democracy_quality"],
    },
    {
        "id": "G007_64",
        "name": "Trust: People in general",
        "description": "WVS longitudinal trust in people in general item",
        "wave": "timeseries",
        "harvest_metric_candidates": ["social_trust"],
    },
    {
        "id": "Y022",
        "name": "Welzel equality sub-index",
        "description": "WVS longitudinal equality composite item",
        "wave": "timeseries",
        "harvest_metric_candidates": ["gender_equality"],
    },
)
_WHO_STATIC_INDICATORS: tuple[dict[str, Any], ...] = (
    # ── Mortality & Life Expectancy ──
    {
        "id": "WHOSIS_000001",
        "IndicatorCode": "WHOSIS_000001",
        "IndicatorName": "Life expectancy at birth (years)",
        "name": "Life expectancy at birth (years)",
        "description": "WHO GHO life expectancy at birth indicator",
        "harvest_metric_candidates": ["health_outcomes", "life_expectancy"],
    },
    {
        "id": "WHOSIS_000002",
        "IndicatorCode": "WHOSIS_000002",
        "IndicatorName": "Healthy life expectancy (HALE) at birth (years)",
        "name": "Healthy life expectancy (HALE) at birth (years)",
        "description": "WHO GHO healthy life expectancy at birth indicator",
        "harvest_metric_candidates": ["health_outcomes", "life_expectancy"],
    },
    {
        "id": "WHOSIS_000015",
        "IndicatorCode": "WHOSIS_000015",
        "IndicatorName": "Life expectancy at age 60 (years)",
        "name": "Life expectancy at age 60 (years)",
        "description": "WHO GHO life expectancy at age 60 indicator",
        "harvest_metric_candidates": ["health_outcomes", "life_expectancy"],
    },
    {
        "id": "WHOSIS_000003",
        "IndicatorCode": "WHOSIS_000003",
        "IndicatorName": "Neonatal mortality rate (per 1000 live births)",
        "name": "Neonatal mortality rate (per 1000 live births)",
        "description": "WHO GHO neonatal mortality rate",
        "harvest_metric_candidates": ["neonatal_mortality", "infant_mortality"],
    },
    {
        "id": "WHS7_104",
        "IndicatorCode": "WHS7_104",
        "IndicatorName": "Infant mortality rate (per 1000 live births)",
        "name": "Infant mortality rate (per 1000 live births)",
        "description": "WHO GHO infant mortality rate",
        "harvest_metric_candidates": ["infant_mortality", "health_outcomes"],
    },
    {
        "id": "MDG_0000000001",
        "IndicatorCode": "MDG_0000000001",
        "IndicatorName": "Under-five mortality rate (per 1000 live births)",
        "name": "Under-five mortality rate (per 1000 live births)",
        "description": "WHO GHO under-five mortality rate",
        "harvest_metric_candidates": ["infant_mortality", "health_outcomes"],
    },
    {
        "id": "MDG_0000000003",
        "IndicatorCode": "MDG_0000000003",
        "IndicatorName": "Neonatal mortality rate (per 1000 live births)",
        "name": "Neonatal mortality rate (per 1000 live births)",
        "description": "WHO GHO neonatal mortality rate (MDG series)",
        "harvest_metric_candidates": ["neonatal_mortality"],
    },
    {
        "id": "WHS9_95",
        "IndicatorCode": "WHS9_95",
        "IndicatorName": "Maternal mortality ratio (per 100 000 live births)",
        "name": "Maternal mortality ratio (per 100 000 live births)",
        "description": "WHO GHO maternal mortality ratio",
        "harvest_metric_candidates": ["maternal_mortality", "health_outcomes"],
    },
    # ── Infectious Disease ──
    {
        "id": "MDG_0000000020",
        "IndicatorCode": "MDG_0000000020",
        "IndicatorName": "Tuberculosis incidence (per 100 000 population per year)",
        "name": "Tuberculosis incidence (per 100 000 population)",
        "description": "WHO GHO tuberculosis incidence rate",
        "harvest_metric_candidates": ["tuberculosis_incidence"],
    },
    {
        "id": "MALARIA_EST_INCIDENCE",
        "IndicatorCode": "MALARIA_EST_INCIDENCE",
        "IndicatorName": "Estimated malaria incidence (per 1000 population at risk)",
        "name": "Estimated malaria incidence (per 1000 population at risk)",
        "description": "WHO GHO malaria incidence estimate",
        "harvest_metric_candidates": ["malaria_incidence"],
    },
    {
        "id": "HIV_0000000001",
        "IndicatorCode": "HIV_0000000001",
        "IndicatorName": "Estimated number of people (all ages) newly infected with HIV",
        "name": "New HIV infections",
        "description": "WHO GHO new HIV infections estimate",
        "harvest_metric_candidates": ["hiv_prevalence"],
    },
    {
        "id": "WHS3_62",
        "IndicatorCode": "WHS3_62",
        "IndicatorName": "Hepatitis B surface antigen prevalence among children under 5",
        "name": "Hepatitis B prevalence children under 5",
        "description": "WHO GHO hepatitis B prevalence",
        "harvest_metric_candidates": ["health_outcomes"],
    },
    # ── NCD / Risk Factors ──
    {
        "id": "NCD_BMI_30A",
        "IndicatorCode": "NCD_BMI_30A",
        "IndicatorName": "Prevalence of obesity among adults, BMI >= 30 (age-standardized estimate) (%)",
        "name": "Prevalence of obesity among adults (BMI >= 30)",
        "description": "WHO GHO adult obesity prevalence",
        "harvest_metric_candidates": ["obesity_prevalence", "health_outcomes"],
    },
    {
        "id": "NCD_CCS_Diab",
        "IndicatorCode": "NCD_CCS_Diab",
        "IndicatorName": "Diabetes prevalence",
        "name": "Diabetes prevalence",
        "description": "WHO GHO diabetes prevalence",
        "harvest_metric_candidates": ["diabetes_prevalence", "health_outcomes"],
    },
    {
        "id": "NCD_HYP_PREVALENCE_A",
        "IndicatorCode": "NCD_HYP_PREVALENCE_A",
        "IndicatorName": "Raised blood pressure (SBP>=140 OR DBP>=90) (age-standardized estimate)",
        "name": "Hypertension prevalence",
        "description": "WHO GHO hypertension prevalence",
        "harvest_metric_candidates": ["hypertension_prevalence", "health_outcomes"],
    },
    {
        "id": "NCDMORT3070",
        "IndicatorCode": "NCDMORT3070",
        "IndicatorName": "Probability (%) of dying between age 30 and exact age 70 from any of cardiovascular disease, cancer, diabetes, or chronic respiratory disease",
        "name": "NCD mortality probability 30-70",
        "description": "WHO GHO NCD premature mortality probability",
        "harvest_metric_candidates": ["noncommunicable_disease_mortality", "health_outcomes"],
    },
    {
        "id": "NCD_TOB_SMOK_CURRE",
        "IndicatorCode": "NCD_TOB_SMOK_CURRE",
        "IndicatorName": "Estimate of current tobacco smoking prevalence (%)",
        "name": "Current tobacco smoking prevalence",
        "description": "WHO GHO tobacco smoking prevalence",
        "harvest_metric_candidates": ["smoking_prevalence"],
    },
    {
        "id": "SA_0000001688",
        "IndicatorCode": "SA_0000001688",
        "IndicatorName": "Total alcohol per capita (>=15) consumption, in litres of pure alcohol",
        "name": "Total alcohol per capita consumption",
        "description": "WHO GHO alcohol consumption per capita",
        "harvest_metric_candidates": ["alcohol_consumption"],
    },
    # ── Health System ──
    {
        "id": "UHC_INDEX_REPORTED",
        "IndicatorCode": "UHC_INDEX_REPORTED",
        "IndicatorName": "UHC index of service coverage",
        "name": "UHC service coverage index",
        "description": "WHO GHO universal health coverage index",
        "harvest_metric_candidates": ["universal_health_coverage", "health_outcomes"],
    },
    {
        "id": "HWF_0001",
        "IndicatorCode": "HWF_0001",
        "IndicatorName": "Medical doctors (per 10 000 population)",
        "name": "Medical doctors per 10 000 population",
        "description": "WHO GHO physician density",
        "harvest_metric_candidates": ["physician_density"],
    },
    {
        "id": "HWF_0006",
        "IndicatorCode": "HWF_0006",
        "IndicatorName": "Hospital beds (per 10 000 population)",
        "name": "Hospital beds per 10 000 population",
        "description": "WHO GHO hospital bed density",
        "harvest_metric_candidates": ["hospital_beds"],
    },
    {
        "id": "WHS6_102",
        "IndicatorCode": "WHS6_102",
        "IndicatorName": "Diphtheria tetanus toxoid and pertussis (DTP3) immunization coverage among 1-year-olds (%)",
        "name": "DTP3 immunization coverage",
        "description": "WHO GHO DTP3 immunization",
        "harvest_metric_candidates": ["vaccination_coverage"],
    },
    {
        "id": "WHS4_117",
        "IndicatorCode": "WHS4_117",
        "IndicatorName": "Measles-containing-vaccine first-dose (MCV1) immunization coverage among 1-year-olds (%)",
        "name": "Measles immunization coverage",
        "description": "WHO GHO measles immunization",
        "harvest_metric_candidates": ["vaccination_coverage"],
    },
    {
        "id": "WHS4_128",
        "IndicatorCode": "WHS4_128",
        "IndicatorName": "Polio (Pol3) immunization coverage among 1-year-olds (%)",
        "name": "Polio immunization coverage",
        "description": "WHO GHO polio immunization",
        "harvest_metric_candidates": ["vaccination_coverage"],
    },
    # ── WASH / Environment ──
    {
        "id": "WSH_WATER_SAFELY_MANAGED",
        "IndicatorCode": "WSH_WATER_SAFELY_MANAGED",
        "IndicatorName": "Population using safely managed drinking-water services (%)",
        "name": "Safely managed drinking water",
        "description": "WHO GHO safely managed water services",
        "harvest_metric_candidates": ["clean_water_access"],
    },
    {
        "id": "WSH_SANITATION_SAFELY_MANAGED",
        "IndicatorCode": "WSH_SANITATION_SAFELY_MANAGED",
        "IndicatorName": "Population using safely managed sanitation services (%)",
        "name": "Safely managed sanitation",
        "description": "WHO GHO safely managed sanitation services",
        "harvest_metric_candidates": ["sanitation_coverage"],
    },
    {
        "id": "SDGAIRBOD_3",
        "IndicatorCode": "SDGAIRBOD_3",
        "IndicatorName": "Ambient and household air pollution attributable death rate (per 100 000 population, age-standardized)",
        "name": "Air pollution attributable death rate",
        "description": "WHO GHO air pollution mortality",
        "harvest_metric_candidates": ["air_pollution_health", "air_quality_index"],
    },
    # ── Reproductive / Child Nutrition ──
    {
        "id": "NUTRITION_HA_2",
        "IndicatorCode": "NUTRITION_HA_2",
        "IndicatorName": "Children aged < 5 years stunted (%)",
        "name": "Child stunting prevalence",
        "description": "WHO GHO child stunting",
        "harvest_metric_candidates": ["child_stunting"],
    },
    {
        "id": "NUTRITION_WH_2",
        "IndicatorCode": "NUTRITION_WH_2",
        "IndicatorName": "Children aged < 5 years wasted (%)",
        "name": "Child wasting prevalence",
        "description": "WHO GHO child wasting",
        "harvest_metric_candidates": ["child_stunting"],
    },
    {
        "id": "NUTRITION_WA_2",
        "IndicatorCode": "NUTRITION_WA_2",
        "IndicatorName": "Children aged < 5 years underweight (%)",
        "name": "Child underweight prevalence",
        "description": "WHO GHO child underweight",
        "harvest_metric_candidates": ["child_stunting"],
    },
    {
        "id": "NUTRITION_ANE_WRA_P",
        "IndicatorCode": "NUTRITION_ANE_WRA_P",
        "IndicatorName": "Anaemia prevalence in women of reproductive age (%)",
        "name": "Anaemia prevalence in women",
        "description": "WHO GHO anaemia prevalence women reproductive age",
        "harvest_metric_candidates": ["health_outcomes"],
    },
    # ── Violence / Injury ──
    {
        "id": "VIOLENCE_HOMICIDERATE",
        "IndicatorCode": "VIOLENCE_HOMICIDERATE",
        "IndicatorName": "Estimates of rates of homicides per 100 000 population",
        "name": "Homicide rate",
        "description": "WHO GHO homicide rate",
        "harvest_metric_candidates": ["homicide_rate"],
    },
    {
        "id": "VIOLENCE_YPLLRATE",
        "IndicatorCode": "VIOLENCE_YPLLRATE",
        "IndicatorName": "Years of potential life lost from violence",
        "name": "Years of life lost from violence",
        "description": "WHO GHO violence years of life lost",
        "harvest_metric_candidates": ["homicide_rate"],
    },
    # ── Mental Health ──
    {
        "id": "MH_12",
        "IndicatorCode": "MH_12",
        "IndicatorName": "Crude suicide rates (per 100 000 population)",
        "name": "Suicide rate",
        "description": "WHO GHO crude suicide rate",
        "harvest_metric_candidates": ["suicide_rate"],
    },
    # ── Road Safety ──
    {
        "id": "RS_198",
        "IndicatorCode": "RS_198",
        "IndicatorName": "Estimated road traffic death rate (per 100 000 population)",
        "name": "Road traffic death rate",
        "description": "WHO GHO road traffic mortality",
        "harvest_metric_candidates": ["health_outcomes"],
    },
    # ── Reproductive Health ──
    {
        "id": "MDG_0000000025",
        "IndicatorCode": "MDG_0000000025",
        "IndicatorName": "Adolescent birth rate (per 1000 women aged 15-19 years)",
        "name": "Adolescent birth rate",
        "description": "WHO GHO adolescent fertility",
        "harvest_metric_candidates": ["fertility_rate", "health_outcomes"],
    },
    {
        "id": "MDG_0000000026",
        "IndicatorCode": "MDG_0000000026",
        "IndicatorName": "Births attended by skilled health personnel (%)",
        "name": "Skilled birth attendance",
        "description": "WHO GHO skilled birth attendance",
        "harvest_metric_candidates": ["maternal_mortality", "health_outcomes"],
    },
    # ── Expenditure ──
    {
        "id": "GHED_CHE_pc_PPP_SHA2011",
        "IndicatorCode": "GHED_CHE_pc_PPP_SHA2011",
        "IndicatorName": "Current health expenditure (CHE) per capita in PPP int$",
        "name": "Health expenditure per capita PPP",
        "description": "WHO GHO health expenditure per capita",
        "harvest_metric_candidates": ["health_spending"],
    },
    {
        "id": "GHED_OOPS_SHA2011",
        "IndicatorCode": "GHED_OOPS_SHA2011",
        "IndicatorName": "Out-of-pocket spending as percentage of current health expenditure (CHE) (%)",
        "name": "Out-of-pocket health spending %",
        "description": "WHO GHO out-of-pocket health expenditure",
        "harvest_metric_candidates": ["out_of_pocket_spending", "health_spending"],
    },
)
_WVS_LOCAL_METRIC_CANDIDATES: dict[str, tuple[str, ...]] = {
    str(item["id"]).strip().upper(): tuple(
        str(metric) for metric in item.get("harvest_metric_candidates", []) if str(metric).strip()
    )
    for item in _WVS_STATIC_INDICATORS
}
_WHO_STATIC_INDICATOR_NAMES: dict[str, str] = {
    str(item["id"]).strip().upper(): str(item["name"]) for item in _WHO_STATIC_INDICATORS
}
_SPARQL_SOURCE_TEMPLATES: dict[str, tuple[dict[str, Any], ...]] = {
    "wikidata_sparql": (
        {
            "id": "country_entities",
            "title": "Wikidata country entities for comparator alignment",
            "description": "Entity resolution for countries, regions, and comparator panels.",
            "keywords": [
                "entity resolution",
                "country",
                "region",
                "wikidata",
                "migration",
                "population",
            ],
            "format": "JSON",
        },
        {
            "id": "indicator_topics",
            "title": "Wikidata policy indicator topics",
            "description": "Ontology enrichment for policy topics, metrics, and indicator concepts.",
            "keywords": [
                "indicator",
                "taxonomy",
                "policy",
                "concept",
                "gdp",
                "unemployment",
                "inflation",
                "health",
                "education",
            ],
            "format": "JSON",
        },
    ),
    "dbpedia_sparql": (
        {
            "id": "country_labels",
            "title": "DBpedia country labels and aliases",
            "description": "Alternative labels and aliases for country/entity matching.",
            "keywords": [
                "alias",
                "labels",
                "country",
                "entity resolution",
                "migration",
                "population",
            ],
            "format": "JSON",
        },
        {
            "id": "policy_taxonomy",
            "title": "DBpedia policy taxonomy hints",
            "description": "Taxonomy enrichment for policy domains and linked concepts.",
            "keywords": ["taxonomy", "policy", "concept", "dbpedia", "gdp", "health", "education"],
            "format": "JSON",
        },
    ),
}
_REST_SOURCE_TEMPLATES: dict[str, tuple[dict[str, Any], ...]] = {
    "openaq_v2": (
        {
            "id": "openaq_air_quality_city_day",
            "title": "OpenAQ city-day air quality aggregates",
            "description": "Rolling-window air quality measurements for UA, PL, RO, MD, and DE city-day comparator panels.",
            "keywords": ["air quality", "pollution", "environment", "city", "daily"],
            "formats": ["JSON"],
            "default_filters": {"country": ["UA", "PL", "RO", "MD", "DE"]},
        },
    ),
    "open_meteo": (
        {
            "id": "open_meteo_country_daily",
            "title": "Open-Meteo country daily weather aggregates",
            "description": "Rolling-window daily weather covariates for canonical country coordinates.",
            "keywords": ["weather", "temperature", "precipitation", "daily"],
            "formats": ["JSON"],
        },
    ),
    "eia_api": (
        {
            "id": "eia_energy_monthly",
            "title": "EIA monthly energy series",
            "description": "Monthly energy indicators and prices for compact policy covariates.",
            "keywords": ["energy", "electricity", "gas", "monthly"],
            "formats": ["JSON"],
        },
    ),
}


def _utc_slug() -> str:
    return datetime.now(UTC).strftime("%Y%m%dT%H%M%SZ")


def _latest_snapshot_dir(source_root: Path) -> Path | None:
    if not source_root.exists():
        return None
    dirs = sorted([p for p in source_root.iterdir() if p.is_dir()])
    return dirs[-1] if dirs else None


def _wvs_raw_dir() -> Path:
    return Path(__file__).resolve().parents[4] / "data" / "raw" / "wvs"


def _wvs_registry_path() -> Path:
    return (
        Path(__file__).resolve().parents[4]
        / "data"
        / "dataset_catalog"
        / "wvs_indicator_registry.yaml"
    )


def _wvs_variable_catalog_path() -> Path:
    return (
        _wvs_raw_dir()
        / "F00003844-WVS_Time_Series_List_of_Variables_and_equivalences_1981_2022_v3_1.xlsx"
    )


@lru_cache(maxsize=1)
def _load_wvs_indicator_registry() -> dict[str, dict[str, Any]]:
    """Load the WVS indicator registry YAML.  Returns ``{code: spec}``."""
    registry_path = _wvs_registry_path()
    if not registry_path.exists():
        return {}
    try:
        import yaml

        data = yaml.safe_load(registry_path.read_text(encoding="utf-8"))
        indicators = data.get("indicators", {}) if isinstance(data, dict) else {}
        return {str(k).strip().upper(): v for k, v in indicators.items() if isinstance(v, dict)}
    except Exception:
        logger.warning(
            "Failed to load WVS indicator registry from {}", registry_path, exc_info=True
        )
        return {}


@lru_cache(maxsize=1)
def _load_wvs_indicator_catalog_from_local_file() -> tuple[dict[str, Any], ...]:
    # First try loading from registry YAML (auto-generated from codebook)
    registry = _load_wvs_indicator_registry()
    if registry:
        catalog: list[dict[str, Any]] = []
        for variable, spec in registry.items():
            title = str(spec.get("title", variable))
            candidates = spec.get("canonical_candidates", [])
            if isinstance(candidates, str):
                candidates = [candidates]
            # Also check the static metric candidates
            static_candidates = _WVS_LOCAL_METRIC_CANDIDATES.get(variable, ())
            merged_candidates = list(dict.fromkeys(list(candidates) + list(static_candidates)))
            item: dict[str, Any] = {
                "id": variable,
                "name": title,
                "description": f"World Values Survey longitudinal variable {variable}: {title}",
                "wave": "timeseries",
            }
            if merged_candidates:
                item["harvest_metric_candidates"] = merged_candidates
            catalog.append(item)
        if catalog:
            logger.info("Loaded {} WVS indicators from registry", len(catalog))
            return tuple(catalog)

    # Fallback: try Excel codebook (read ALL indicators, not just supported)
    path = _wvs_variable_catalog_path()
    if not path.exists():
        return _WVS_STATIC_INDICATORS

    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl is unavailable; falling back to static WVS indicator catalog")
        return _WVS_STATIC_INDICATORS

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        header_index: dict[str, int] = {}
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, ())
        for index, cell in enumerate(header):
            label = str(cell or "").strip()
            if label:
                header_index[label] = index

        variable_idx = header_index.get("Variable")
        title_idx = header_index.get("Title")
        if variable_idx is None or title_idx is None:
            return _WVS_STATIC_INDICATORS

        catalog = []
        for row in rows:
            variable = str(row[variable_idx] or "").strip().upper()
            if not variable:
                continue
            title = str(row[title_idx] or "").strip()
            item: dict[str, Any] = {
                "id": variable,
                "name": title or variable,
                "description": f"World Values Survey longitudinal variable {variable}: {title}",
                "wave": "timeseries",
            }
            metric_candidates = _WVS_LOCAL_METRIC_CANDIDATES.get(variable, ())
            if metric_candidates:
                item["harvest_metric_candidates"] = list(metric_candidates)
            catalog.append(item)
        if catalog:
            logger.info("Loaded {} WVS indicators from Excel codebook", len(catalog))
            return tuple(catalog)
    except Exception:
        logger.warning("Failed to load local WVS indicator catalog from {}", path, exc_info=True)

    return _WVS_STATIC_INDICATORS


def _read_jsonl(path: Path) -> list[dict]:
    rows: list[dict] = []
    if not path.exists():
        return rows
    with open(path, encoding="utf-8") as fh:
        for line in fh:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def _write_jsonl(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as fh:
        for row in rows:
            fh.write(json.dumps(row, ensure_ascii=False) + "\n")


def _apply_limit(rows: list[dict], limit: int) -> list[dict]:
    capped = max(int(limit), 0)
    return rows[:capped] if capped else rows


def _is_preflight_run(config: DatasetBatchConfig) -> bool:
    return bool(config.preflight_only or config.run_profile == "preflight_core")


def _effective_dataset_limit(config: DatasetBatchConfig) -> int:
    configured = max(int(config.max_datasets_per_source), 0)
    if not _is_preflight_run(config):
        return configured
    if configured <= 0 or configured >= 100_000:
        return 50
    return min(configured, 50)


def _harvest_limit_for_source(spec: SourceSpec, config: DatasetBatchConfig) -> int:
    limit = _effective_dataset_limit(config)
    sampled_like = bool(config.is_sampled_run or _is_preflight_run(config))
    if limit <= 0 or not sampled_like:
        return limit
    if _is_preflight_run(config):
        if spec.family == "poland_api":
            return max(limit * 4, 80)
        if spec.family == "ckan":
            return max(limit * 4, 120)
        if spec.family == "worldbank":
            return max(limit * 6, 120)
        if spec.family == "who":
            return max(limit * 6, 120)
        if spec.family in {"uis", "unpd"}:
            return max(limit * 4, 60)
        if spec.family == "wvs":
            return max(limit * 2, 25)
        if spec.family == "opendatasoft":
            return max(limit * 4, 80)
        if spec.family == "socrata":
            return max(limit * 4, 80)
        if spec.family == "sparql":
            return max(limit, 10)
        if spec.family == "rest":
            return max(limit, 5)
        return max(limit * 2, limit)
    if spec.family == "poland_api":
        return max(limit * 150, 1200)
    if spec.family == "ckan":
        return max(limit * 50, 200)
    if spec.family == "worldbank":
        return max(limit * 40, 250)
    if spec.family == "who":
        return max(limit * 50, 250)
    if spec.family in {"uis", "unpd"}:
        return max(limit * 25, 120)
    if spec.family == "wvs":
        return max(limit * 10, 25)
    if spec.family == "opendatasoft":
        return max(limit * 20, 200)
    if spec.family == "socrata":
        return max(limit * 20, 200)
    if spec.family == "sparql":
        return max(limit, 10)
    if spec.family == "rest":
        return max(limit, 5)
    return limit


async def harvest_sources(config: DatasetBatchConfig) -> dict[str, list[dict]]:
    """Harvest all enabled sources in selected wave."""
    registry = config.load_registry()
    specs = registry.enabled_sources(wave=config.wave, run_profile=config.run_profile)
    metrics_map = load_metrics_map(config.resolved_metrics_map_path)
    checkpoint = load_json(config.harvest_checkpoint_path, default={})
    if not isinstance(checkpoint, dict):
        checkpoint = {}

    started_at = datetime.now(UTC).isoformat()
    out: dict[str, list[dict]] = {}
    state_lock = asyncio.Lock()
    semaphore = asyncio.Semaphore(_HARVEST_MAX_PARALLELISM)
    completed_names: set[str] = set()
    pending_specs = list(specs)

    async def _record_success(spec: SourceSpec, rows: list[dict]) -> None:
        payload_path = _current_snapshot_payload_path(config, spec.name)
        async with state_lock:
            out[spec.name] = rows
            checkpoint[spec.name] = {
                "status": "complete",
                "records_fetched": len(rows),
                "row_bytes": int(payload_path.stat().st_size) if payload_path.exists() else 0,
                "cursor": None,
                "offset": 0,
                "page": 0,
                "etag": None,
                "last_modified": None,
                "payload_hash": _payload_hash(payload_path) if payload_path.exists() else "",
                "last_success_at": datetime.now(UTC).isoformat(),
                "error": "",
            }
            completed_names.add(spec.name)
            write_json(config.harvest_checkpoint_path, checkpoint)

    async def _record_failure(spec: SourceSpec, exc: Exception) -> None:
        logger.error("Harvest failed for source {}: {}", spec.name, exc)
        async with state_lock:
            out[spec.name] = []
            checkpoint[spec.name] = {
                "status": "failed",
                "records_fetched": 0,
                "row_bytes": 0,
                "cursor": None,
                "offset": 0,
                "page": 0,
                "etag": None,
                "last_modified": None,
                "payload_hash": "",
                "last_success_at": "",
                "error": str(exc)[:500],
            }
            completed_names.add(spec.name)
            write_json(config.harvest_checkpoint_path, checkpoint)

    async def _harvest_spec(spec: SourceSpec) -> None:
        try:
            rows = await harvest_one_source(
                spec,
                config,
                harvested=out,
                metrics_map=metrics_map,
                checkpoint=checkpoint,
            )
        except Exception as exc:
            await _record_failure(spec, exc)
            return
        await _record_success(spec, rows)

    async def _run_parallel_spec(spec: SourceSpec) -> None:
        async with semaphore:
            await _harvest_spec(spec)

    async def _run_serial_group(group_specs: list[SourceSpec]) -> None:
        for spec in group_specs:
            await _harvest_spec(spec)

    while pending_specs:
        ready_specs = [
            spec
            for spec in pending_specs
            if not spec.seed_from or spec.seed_from in completed_names
        ]
        if not ready_specs:
            unresolved = ", ".join(f"{spec.name}->{spec.seed_from}" for spec in pending_specs)
            raise RuntimeError(f"Harvest dependency deadlock: {unresolved}")

        serial_specs = [spec for spec in ready_specs if _harvest_runs_serially(spec)]
        parallel_specs = [spec for spec in ready_specs if not _harvest_runs_serially(spec)]
        tasks = [asyncio.create_task(_run_parallel_spec(spec)) for spec in parallel_specs]
        if serial_specs:
            tasks.append(asyncio.create_task(_run_serial_group(serial_specs)))
        if tasks:
            await asyncio.gather(*tasks)
        ready_names = {spec.name for spec in ready_specs}
        pending_specs = [spec for spec in pending_specs if spec.name not in ready_names]

    stage_manifest = config.manifests_dir / "harvest.json"
    write_stage_manifest(
        manifest_path=stage_manifest,
        stage="harvest",
        status="ok",
        metrics={
            "wave": config.wave or "ALL",
            "sources": len(specs),
            "records": sum(len(v) for v in out.values()),
        },
        artifacts=[],
        started_at=started_at,
    )
    return out


def _harvest_runs_serially(spec: SourceSpec) -> bool:
    return spec.name in _SERIAL_HARVEST_SOURCES


async def harvest_one_source(
    spec: SourceSpec,
    config: DatasetBatchConfig,
    *,
    harvested: dict[str, list[dict]] | None = None,
    metrics_map: dict[str, dict] | None = None,
    checkpoint: dict[str, Any] | None = None,
) -> list[dict]:
    """Harvest one source with optional resume from latest raw snapshot."""
    source_root = config.raw_dir / spec.name
    latest_dir = _latest_snapshot_dir(source_root)
    latest_payload = latest_dir / "payload.jsonl" if latest_dir else None
    current_snapshot_dir = _current_source_snapshot_dir(config, spec.name)
    current_payload = current_snapshot_dir / "payload.jsonl"
    existing_entry = (checkpoint or {}).get(spec.name) if isinstance(checkpoint, dict) else None
    if (
        config.resume
        and isinstance(existing_entry, dict)
        and str(existing_entry.get("status")) == "complete"
        and current_payload.exists()
    ):
        logger.info("Using in-progress snapshot payload for {}: {}", spec.name, current_payload)
        rows = _read_jsonl(current_payload)
        rows = _prioritize_rows_for_sampling(
            rows, spec=spec, config=config, metrics_map=metrics_map
        )
        return _apply_limit(rows, _effective_dataset_limit(config))

    if config.resume and latest_payload and latest_payload.exists():
        logger.info("Using cached raw snapshot for {}: {}", spec.name, latest_payload)
        rows = _read_jsonl(latest_payload)
        rows = _prioritize_rows_for_sampling(
            rows, spec=spec, config=config, metrics_map=metrics_map
        )
        return _apply_limit(rows, _effective_dataset_limit(config))

    logger.info("Harvesting source {} ({})", spec.name, spec.endpoint)
    harvest_limit = _harvest_limit_for_source(spec, config)
    if spec.seed_from:
        rows = _harvest_from_seed_source(spec, config, harvested=harvested)
    elif spec.family == "ckan":
        rows = await _harvest_ckan(spec.endpoint, harvest_limit, config.harvest_timeout)
    elif spec.family == "poland_api":
        rows = await _harvest_poland_open_data(spec.endpoint, harvest_limit, config.harvest_timeout)
    elif spec.family == "opendatasoft":
        rows = await _harvest_opendatasoft(spec.endpoint, harvest_limit, config.harvest_timeout)
    elif spec.family == "socrata":
        rows = await _harvest_socrata(spec.endpoint, harvest_limit, config.harvest_timeout)
    elif spec.family == "sparql":
        rows = _harvest_sparql_templates(spec)
    elif spec.family == "rest":
        rows = _harvest_rest_catalog(spec)
    elif spec.family == "worldbank":
        rows = await _harvest_worldbank(spec.endpoint, harvest_limit, config.harvest_timeout)
        rows = _augment_worldbank_rows(rows, metrics_map=metrics_map)
    elif spec.family == "ukons":
        rows = await _harvest_ukons(spec.endpoint, harvest_limit, config.harvest_timeout)
    elif spec.family == "undata":
        rows = await _harvest_undata(spec, config.harvest_timeout)
    elif spec.family == "sdmx":
        rows = await _harvest_sdmx_dataflows(spec, config.harvest_timeout)
    elif spec.family == "who":
        rows = await _harvest_who_indicators(spec.endpoint, harvest_limit, config.harvest_timeout)
        rows = _augment_who_rows(rows, metrics_map=metrics_map)
    elif spec.family == "uis":
        rows = await _harvest_uis_indicators(spec.endpoint, harvest_limit, config.harvest_timeout)
    elif spec.family == "wvs":
        rows = await _harvest_wvs(spec.endpoint, harvest_limit, config.harvest_timeout)
    elif spec.family == "unpd":
        rows = await _harvest_unpd_indicators(spec.endpoint, harvest_limit, config.harvest_timeout)
    else:
        logger.warning("Unknown source family '{}' for {}", spec.family, spec.name)
        rows = []
    rows = _prioritize_rows_for_sampling(rows, spec=spec, config=config, metrics_map=metrics_map)
    rows = _apply_limit(rows, _effective_dataset_limit(config))

    ts_dir = current_snapshot_dir
    payload_path = ts_dir / "payload.jsonl"
    manifest_path = ts_dir / "manifest.json"
    _write_jsonl(payload_path, rows)

    write_raw_manifest(
        manifest_path=manifest_path,
        source=spec.name,
        endpoint=spec.endpoint,
        payload_path=payload_path,
        count=len(rows),
        filters={
            "agency_prefix": spec.agency_prefix,
            "agency_allowlist": list(spec.agency_allowlist),
            "exclude_agencies": list(spec.exclude_agencies),
            "wave": spec.wave,
            "run_lane": spec.run_lane,
            "publish_blocking": spec.publish_blocking,
            "history_policy": spec.history_policy,
            "default_lookback_days": spec.default_lookback_days,
            "max_rows_per_snapshot": spec.max_rows_per_snapshot,
            "max_bytes_per_snapshot": spec.max_bytes_per_snapshot,
            "allow_manual_backfill": spec.allow_manual_backfill,
            "seed_from": spec.seed_from,
            "format_allowlist": list(spec.format_allowlist),
        },
        parser_version="2",
    )
    return rows


def _current_source_snapshot_dir(config: DatasetBatchConfig, source_name: str) -> Path:
    return config.raw_dir / source_name / config.snapshot_root.name


def _current_snapshot_payload_path(config: DatasetBatchConfig, source_name: str) -> Path:
    return _current_source_snapshot_dir(config, source_name) / "payload.jsonl"


def _payload_hash(path: Path) -> str:
    if not path.exists():
        return ""
    stat = path.stat()
    return hash_payload(
        {"path": str(path), "size": int(stat.st_size), "mtime_ns": int(stat.st_mtime_ns)}
    )


def _flatten_text_values(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, str):
        text = value.strip()
        return [text] if text else []
    if isinstance(value, dict):
        out: list[str] = []
        for nested in value.values():
            out.extend(_flatten_text_values(nested))
        return out
    if isinstance(value, list):
        out: list[str] = []
        for item in value:
            out.extend(_flatten_text_values(item))
        return out
    text = str(value).strip()
    return [text] if text else []


def _row_text(row: dict[str, Any]) -> str:
    body_parts: list[str] = []
    for key in (
        "name",
        "title",
        "description",
        "Definition",
        "sourceNote",
        "notes",
        "keywords",
        "category",
        "categories",
        "regions",
        "themes",
    ):
        body_parts.extend(_flatten_text_values(row.get(key)))
    tags = " ".join(_flatten_text_values(row.get("tags")))
    body = " ".join(body_parts)
    return f"{body} {tags}".strip().lower()


def _priority_metric_hits(text: str) -> list[str]:
    hits: list[str] = []
    for metric_id in _PRIORITY_PATTERNS:
        patterns = _PRIORITY_PATTERNS.get(metric_id, ())
        if any(pattern in text for pattern in patterns):
            hits.append(metric_id)
    return hits


def _score_row_for_sampling(
    row: dict[str, Any],
    *,
    spec: SourceSpec,
    metrics_map: dict[str, dict] | None,
) -> tuple[int, list[str]]:
    text = _row_text(row)
    matched_metrics = map_to_polisyos_metrics(row, metrics_map)
    heuristic_hits = _priority_metric_hits(text)
    candidates: list[str] = []
    for metric_id in [*matched_metrics, *heuristic_hits]:
        if metric_id not in candidates:
            candidates.append(metric_id)

    priority_hits = [metric for metric in candidates if metric in _SMOKE_PRIORITY_METRICS]
    score = 0
    score += 100 * len(priority_hits)
    score += 25 * len(candidates)
    score += 4 * sum(
        1
        for metric_id in _SMOKE_PRIORITY_METRICS
        if any(pattern in text for pattern in _PRIORITY_PATTERNS.get(metric_id, ()))
    )

    if spec.family == "sdmx":
        code = str(row.get("id", "") or "").lower()
        if any(
            token in code
            for token in ("une", "unemp", "lfs", "cpi", "gdp", "pov", "mig", "edu", "health")
        ):
            score += 10

    if any(token in text for token in _PRIORITY_BLACKLIST):
        score -= 15
    if row.get("description"):
        score += 2
    return score, candidates


def _prioritize_rows_for_sampling(
    rows: list[dict],
    *,
    spec: SourceSpec,
    config: DatasetBatchConfig,
    metrics_map: dict[str, dict] | None,
) -> list[dict]:
    sampled_like = bool(config.is_sampled_run or _is_preflight_run(config))
    if not rows or not sampled_like:
        return rows
    if spec.family not in {"sdmx", "who", "uis", "unpd", "worldbank", "wvs", "ckan", "poland_api"}:
        return rows

    scored: list[tuple[int, int, tuple[str, ...], dict]] = []
    for index, row in enumerate(rows):
        score, candidates = _score_row_for_sampling(row, spec=spec, metrics_map=metrics_map)
        if candidates:
            row = dict(row)
            row["harvest_metric_candidates"] = candidates
        scored.append((score, index, tuple(candidates), row))

    ranked = sorted(
        scored,
        key=lambda item: (
            -int(bool(item[2])) if spec.metrics_required else 0,
            -item[0],
            item[1],
        ),
    )
    selected_indexes: list[int] = []
    selected_set: set[int] = set()
    limit = _effective_dataset_limit(config)

    if limit > 0:
        priority_order = _SOURCE_PRIORITY_METRICS.get(spec.name, _SMOKE_PRIORITY_METRICS)
        for metric_id in priority_order:
            for _score, index, candidates, _row in ranked:
                if metric_id not in candidates or index in selected_set:
                    continue
                selected_indexes.append(index)
                selected_set.add(index)
                break
            if len(selected_indexes) >= limit:
                break

        if len(selected_indexes) < limit:
            candidate_ranked = [item for item in ranked if item[2]]
            fallback_ranked = (
                candidate_ranked if spec.metrics_required and candidate_ranked else ranked
            )
            for _score, index, _candidates, _row in fallback_ranked:
                if index in selected_set:
                    continue
                selected_indexes.append(index)
                selected_set.add(index)
                if len(selected_indexes) >= limit:
                    break

    ranked_by_index = {index: row for _score, index, _candidates, row in ranked}
    prioritized = [ranked_by_index[index] for index in selected_indexes if index in ranked_by_index]
    prioritized.extend(
        row for _score, index, _candidates, row in ranked if index not in selected_set
    )
    logger.info(
        "Prioritized {} sampled rows for {} before applying limit {}",
        len(prioritized),
        spec.name,
        limit,
    )
    return prioritized


def _augment_worldbank_rows(
    rows: list[dict],
    *,
    metrics_map: dict[str, dict] | None,
) -> list[dict]:
    if not metrics_map:
        return rows
    seen = {str(row.get("id") or "").strip().upper() for row in rows if isinstance(row, dict)}
    augmented = list(rows)
    for metric_id, spec in metrics_map.items():
        if not isinstance(spec, dict):
            continue
        keywords = [str(value).strip() for value in spec.get("keywords", []) if str(value).strip()]
        for indicator_id in spec.get("worldbank_indicators", []) or []:
            code = str(indicator_id or "").strip()
            if not code:
                continue
            upper_code = code.upper()
            if upper_code in seen:
                continue
            seen.add(upper_code)
            title = keywords[0] if keywords else code
            augmented.append(
                {
                    "id": code,
                    "name": title,
                    "sourceNote": " ".join(keywords),
                    "harvest_metric_candidates": [metric_id],
                }
            )
    return augmented


def _augment_who_rows(
    rows: list[dict],
    *,
    metrics_map: dict[str, dict] | None,
) -> list[dict]:
    seen = {
        str(row.get("IndicatorCode") or row.get("id") or "").strip().upper()
        for row in rows
        if isinstance(row, dict)
    }
    augmented = list(rows)
    planned: list[tuple[str, list[str], str]] = []
    if metrics_map:
        for metric_id, spec in metrics_map.items():
            if not isinstance(spec, dict):
                continue
            keywords = [
                str(value).strip() for value in spec.get("keywords", []) if str(value).strip()
            ]
            for indicator_id in spec.get("who_indicators", []) or []:
                code = str(indicator_id or "").strip().upper()
                if code:
                    planned.append((code, [metric_id], keywords[0] if keywords else code))
    if not planned:
        for item in _WHO_STATIC_INDICATORS:
            code = str(item.get("id") or "").strip().upper()
            metrics = [
                str(metric_id).strip()
                for metric_id in item.get("harvest_metric_candidates", [])
                if str(metric_id).strip()
            ]
            planned.append((code, metrics, str(item.get("name") or code)))
    for code, metric_candidates, title in planned:
        if code in seen:
            continue
        seen.add(code)
        augmented.append(
            {
                "id": code,
                "IndicatorCode": code,
                "IndicatorName": _WHO_STATIC_INDICATOR_NAMES.get(code, title),
                "name": _WHO_STATIC_INDICATOR_NAMES.get(code, title),
                "description": f"WHO GHO curated indicator {code}",
                "harvest_metric_candidates": metric_candidates,
            }
        )
    return augmented


def _harvest_from_seed_source(
    spec: SourceSpec,
    config: DatasetBatchConfig,
    *,
    harvested: dict[str, list[dict]] | None = None,
) -> list[dict]:
    rows = list((harvested or {}).get(spec.seed_from, []))
    if not rows:
        seed_root = config.raw_dir / spec.seed_from
        latest_dir = _latest_snapshot_dir(seed_root)
        latest_payload = latest_dir / "payload.jsonl" if latest_dir else None
        rows = _read_jsonl(latest_payload) if latest_payload else []

    curated: list[dict] = []
    for row in rows:
        candidate = _curate_seed_row(row, spec)
        if candidate is not None:
            curated.append(row)
            if candidate is not row:
                curated[-1] = candidate

    limit = max(int(config.max_datasets_per_source), 0)
    return curated[:limit] if limit else curated


def _curate_seed_row(row: dict[str, Any], spec: SourceSpec) -> dict[str, Any] | None:
    if spec.family == "ckan":
        return curate_ckan_package(row, spec)
    if not _generic_seed_row_allowed(row, spec):
        return None
    return dict(row)


def _generic_seed_row_allowed(row: dict[str, Any], spec: SourceSpec) -> bool:
    text = _row_text(row)
    allow_match = (
        _matches_any_seed(text, spec.keyword_allowlist) if spec.keyword_allowlist else True
    )
    deny_match = _matches_any_seed(text, spec.keyword_denylist)
    if not allow_match or deny_match:
        return False

    formats = _seed_row_formats(row)
    if spec.family == "poland_api" and row.get("resources_related_url"):
        # Poland catalog rows often advertise only the landing-page HTML format even when
        # the related resources API resolves to machine-readable JSON distributions later.
        if not any(fmt for fmt in formats if fmt != "HTML") and spec.format_allowlist:
            formats = tuple(spec.format_allowlist)
    if spec.format_denylist and any(fmt in spec.format_denylist for fmt in formats):
        return False
    if spec.format_allowlist and not any(fmt in spec.format_allowlist for fmt in formats):
        return False
    if spec.require_curated_resources:
        return bool(
            row.get("resource_url")
            or row.get("resources_related_url")
            or row.get("dataset_url")
            or row.get("url")
        )
    return True


def _seed_row_formats(row: dict[str, Any]) -> tuple[str, ...]:
    formats = [
        str(item).strip().upper() for item in (row.get("formats") or []) if str(item).strip()
    ]
    explicit = str(row.get("format") or "").strip().upper()
    if explicit and explicit not in formats:
        formats.append(explicit)
    return tuple(formats)


def _matches_any_seed(text: str, needles: tuple[str, ...]) -> bool:
    haystack = text.lower()
    for needle in needles:
        candidate = str(needle or "").strip().lower()
        if candidate and candidate in haystack:
            return True
    return False


async def _harvest_ckan(endpoint: str, limit: int, timeout_s: int) -> list[dict]:
    rows: list[dict] = []
    start = 0
    per_page = 100
    min_per_page = 10
    timeout = aiohttp.ClientTimeout(
        total=max(timeout_s * 3, 180),
        sock_connect=max(timeout_s, 30),
        sock_read=max(timeout_s * 2, 120),
    )
    async with aiohttp.ClientSession(timeout=timeout) as session:
        while len(rows) < limit:
            page_size = per_page
            data: dict[str, Any] | list[Any] | None = None
            while data is None:
                params = {"rows": page_size, "start": start, "include_private": "false"}
                try:
                    data = await asyncio.wait_for(
                        _harvest_json_with_retries(
                            session,
                            endpoint,
                            params=params,
                            context=f"ckan start={start} rows={page_size}",
                        ),
                        timeout=max(timeout_s * 2, 90),
                    )
                except TimeoutError:
                    if page_size <= min_per_page:
                        raise
                    next_page_size = max(page_size // 2, min_per_page)
                    logger.warning(
                        "CKAN harvest timed out for {} start={} rows={}; retrying with rows={}",
                        endpoint,
                        start,
                        page_size,
                        next_page_size,
                    )
                    page_size = next_page_size
            per_page = page_size
            if not isinstance(data, dict):
                break
            result = data.get("result", {}) if isinstance(data, dict) else {}
            batch = result.get("results", []) if isinstance(result, dict) else []
            if not batch:
                break
            rows.extend([r for r in batch if isinstance(r, dict)])
            total = int(result.get("count", 0))
            start += per_page
            if start >= min(total, limit):
                break
    return rows[:limit]


async def _harvest_json_with_retries(
    session: aiohttp.ClientSession,
    url: str,
    *,
    params: dict[str, Any],
    context: str,
    headers: dict[str, str] | None = None,
    max_attempts: int = 3,
) -> dict[str, Any] | list[Any] | None:
    attempt = 0
    request_headers = {"Accept": "application/json", "User-Agent": "PolicyOS/1.0 datasets-batch"}
    if headers:
        request_headers.update(headers)
    while True:
        try:
            async with session.get(url, params=params, headers=request_headers) as resp:
                if resp.status == 200:
                    return await resp.json(content_type=None)
                if (
                    resp.status in {408, 425, 429, 500, 502, 503, 504}
                    and attempt < max_attempts - 1
                ):
                    delay = _harvest_retry_delay_seconds(resp.headers.get("Retry-After"), attempt)
                    logger.warning(
                        "Retrying harvest request for {} after HTTP {} (sleep={}s)",
                        context,
                        resp.status,
                        delay,
                    )
                    await asyncio.sleep(delay)
                    attempt += 1
                    continue
                logger.warning("Harvest request failed for {} with HTTP {}", context, resp.status)
                return None
        except (TimeoutError, aiohttp.ClientError) as exc:
            if attempt >= max_attempts - 1:
                raise
            delay = _harvest_retry_delay_seconds(None, attempt)
            logger.warning(
                "Retrying harvest request for {} after transient error (sleep={}s): {}",
                context,
                delay,
                exc,
            )
            await asyncio.sleep(delay)
            attempt += 1


def _harvest_retry_delay_seconds(retry_after: str | None, attempt: int) -> float:
    if retry_after:
        try:
            return min(max(float(retry_after), 1.0), 60.0)
        except ValueError:
            pass
    return min(5.0 * (2**attempt), 30.0)


async def _harvest_opendatasoft(endpoint: str, limit: int, timeout_s: int) -> list[dict]:
    rows: list[dict] = []
    offset = 0
    per_page = 100
    timeout = aiohttp.ClientTimeout(total=timeout_s)
    base = endpoint.rstrip("/")
    url = f"{base}/api/explore/v2.1/catalog/datasets"
    async with aiohttp.ClientSession(timeout=timeout) as session:
        while len(rows) < limit:
            params = {"limit": min(per_page, limit - len(rows)), "offset": offset}
            async with session.get(
                url, params=params, headers={"Accept": "application/json"}
            ) as resp:
                if resp.status != 200:
                    break
                payload = await resp.json(content_type=None)
            batch = payload.get("results", []) if isinstance(payload, dict) else []
            if not isinstance(batch, list) or not batch:
                break
            for item in batch:
                if not isinstance(item, dict):
                    continue
                metas = item.get("metas") if isinstance(item.get("metas"), dict) else {}
                default = metas.get("default") if isinstance(metas.get("default"), dict) else metas
                dataset_id = str(item.get("dataset_id") or item.get("datasetid") or "").strip()
                title = str(default.get("title") or dataset_id).strip()
                rows.append(
                    {
                        "id": dataset_id,
                        "title": title,
                        "description": str(default.get("description") or "").strip(),
                        "keywords": [
                            str(tag).strip()
                            for tag in (default.get("keyword") or [])
                            if str(tag).strip()
                        ],
                        "category": str(default.get("theme") or "").strip(),
                        "format": "JSON",
                        "formats": ["JSON"],
                        "schema_fields": [
                            str(field.get("name") or "").strip()
                            for field in (item.get("fields") or [])
                            if isinstance(field, dict) and str(field.get("name") or "").strip()
                        ],
                        "dataset_url": f"{base}/api/explore/v2.1/catalog/datasets/{dataset_id}/records",
                        "metadata_modified": default.get("modified"),
                        "publisher": str(default.get("publisher") or "Opendatasoft").strip(),
                    }
                )
            if len(batch) < per_page:
                break
            offset += per_page
    return rows[:limit]


async def _harvest_socrata(endpoint: str, limit: int, timeout_s: int) -> list[dict]:
    rows: list[dict] = []
    page = 1
    per_page = 100
    timeout = aiohttp.ClientTimeout(total=timeout_s)
    base = endpoint.rstrip("/")
    url = f"{base}/api/views"
    async with aiohttp.ClientSession(timeout=timeout) as session:
        while len(rows) < limit:
            params = {"limit": min(per_page, limit - len(rows)), "page": page}
            async with session.get(
                url, params=params, headers={"Accept": "application/json"}
            ) as resp:
                if resp.status != 200:
                    break
                payload = await resp.json(content_type=None)
            if not isinstance(payload, list) or not payload:
                break
            for item in payload:
                if not isinstance(item, dict):
                    continue
                dataset_id = str(item.get("id") or "").strip()
                rows.append(
                    {
                        "id": dataset_id,
                        "title": str(item.get("name") or dataset_id).strip(),
                        "description": str(item.get("description") or "").strip(),
                        "category": str(item.get("category") or "").strip(),
                        "keywords": [
                            str(item.get("category") or "").strip(),
                            str(item.get("attribution") or "").strip(),
                        ],
                        "format": "JSON",
                        "formats": ["JSON"],
                        "dataset_url": f"{base}/resource/{dataset_id}.json",
                        "metadata_modified": item.get("rowsUpdatedAt"),
                        "publisher": str(item.get("attribution") or "Socrata").strip(),
                    }
                )
            if len(payload) < per_page:
                break
            page += 1
    return rows[:limit]


def _harvest_sparql_templates(spec: SourceSpec) -> list[dict]:
    templates = _SPARQL_SOURCE_TEMPLATES.get(spec.name, ())
    return [
        {
            **template,
            "formats": [template.get("format", "JSON")],
            "dataset_url": spec.endpoint,
            "publisher": spec.name,
        }
        for template in templates
    ]


def _harvest_rest_catalog(spec: SourceSpec) -> list[dict]:
    templates = _REST_SOURCE_TEMPLATES.get(spec.name, ())
    return [
        {
            **template,
            "dataset_url": spec.endpoint,
            "publisher": spec.name,
            "metadata_modified": datetime.now(UTC).date().isoformat(),
        }
        for template in templates
    ]


def _strip_html(text: str) -> str:
    return re.sub(r"<[^>]+>", " ", text or "").replace("&nbsp;", " ").strip()


def _flatten_poland_open_data_row(raw: dict[str, Any]) -> dict[str, Any]:
    attrs = raw.get("attributes") if isinstance(raw.get("attributes"), dict) else {}
    relationships = raw.get("relationships") if isinstance(raw.get("relationships"), dict) else {}
    keywords = [str(item).strip() for item in (attrs.get("keywords") or []) if str(item).strip()]
    category = attrs.get("category") if isinstance(attrs.get("category"), dict) else {}
    category_title = str(category.get("title") or "").strip()
    categories = [
        str(item.get("title") or "").strip()
        for item in (attrs.get("categories") or [])
        if isinstance(item, dict) and str(item.get("title") or "").strip()
    ]
    regions = [
        str(item.get("name") or "").strip()
        for item in (attrs.get("regions") or [])
        if isinstance(item, dict) and str(item.get("name") or "").strip()
    ]
    formats = [
        str(item).strip().upper() for item in (attrs.get("formats") or []) if str(item).strip()
    ]
    types = [str(item).strip() for item in (attrs.get("types") or []) if str(item).strip()]
    institution = (
        relationships.get("institution")
        if isinstance(relationships.get("institution"), dict)
        else {}
    )
    institution_data = institution.get("data") if isinstance(institution.get("data"), dict) else {}
    resources_rel = (
        relationships.get("resources") if isinstance(relationships.get("resources"), dict) else {}
    )
    resources_links = (
        resources_rel.get("links") if isinstance(resources_rel.get("links"), dict) else {}
    )
    resources_related_url = str(resources_links.get("related") or "").strip()
    notes = _strip_html(str(attrs.get("notes") or ""))
    description_parts = [notes, category_title, *categories, *keywords, *regions, *types]
    publisher = str(institution_data.get("id") or "").strip() or "data_gov_pl"
    tag_names = list(dict.fromkeys([*keywords, *categories, category_title]))
    return {
        "id": str(raw.get("id") or ""),
        "name": str(attrs.get("slug") or ""),
        "title": str(attrs.get("title") or attrs.get("slug") or raw.get("id") or ""),
        "notes": notes,
        "description": " ".join(part for part in description_parts if part).strip(),
        "keywords": keywords,
        "category": category_title,
        "categories": categories,
        "regions": regions,
        "types": types,
        "formats": formats,
        "organization": {"name": publisher},
        "tags": [{"name": item} for item in tag_names if item],
        "metadata_modified": attrs.get("resource_modified") or attrs.get("modified"),
        "modified": attrs.get("modified"),
        "resource_modified": attrs.get("resource_modified"),
        "license_name": str(attrs.get("license_condition_original") or ""),
        "resources_related_url": resources_related_url,
        "dataset_url": str(((raw.get("links") or {}).get("self")) or "").strip(),
        "spatial": "POL"
        if any(region.lower() in {"polska", "poland"} for region in regions)
        else "",
    }


async def _harvest_poland_open_data(endpoint: str, limit: int, timeout_s: int) -> list[dict]:
    rows: list[dict] = []
    page = 1
    per_page = 100
    next_url: str | None = None
    timeout = aiohttp.ClientTimeout(total=timeout_s)
    async with aiohttp.ClientSession(timeout=timeout) as session:
        while len(rows) < limit:
            request_url = next_url or endpoint
            params = None if next_url else {"page": page, "per_page": per_page, "sort": "id"}
            async with session.get(
                request_url,
                params=params,
                headers={"Accept": "application/vnd.api+json"},
            ) as resp:
                if resp.status != 200:
                    break
                payload = await resp.json(content_type=None)
            batch = payload.get("data", []) if isinstance(payload, dict) else []
            if not isinstance(batch, list) or not batch:
                break
            rows.extend(
                _flatten_poland_open_data_row(item) for item in batch if isinstance(item, dict)
            )
            links = payload.get("links", {}) if isinstance(payload, dict) else {}
            next_url = str(links.get("next") or "").strip() or None
            page += 1
            if not next_url:
                break
    return rows[:limit]


async def _harvest_worldbank(endpoint: str, limit: int, timeout_s: int) -> list[dict]:
    rows: list[dict] = []
    page = 1
    per_page = 1000
    timeout = aiohttp.ClientTimeout(total=timeout_s)
    async with aiohttp.ClientSession(timeout=timeout) as session:
        while len(rows) < limit:
            params = {"format": "json", "per_page": per_page, "page": page}
            async with session.get(endpoint, params=params) as resp:
                if resp.status != 200:
                    break
                data = await resp.json(content_type=None)
            if not isinstance(data, list) or len(data) < 2:
                break
            batch = data[1] if isinstance(data[1], list) else []
            if not batch:
                break
            rows.extend([r for r in batch if isinstance(r, dict)])
            total = int(data[0].get("total", 0)) if isinstance(data[0], dict) else 0
            page += 1
            if len(rows) >= min(limit, total or limit):
                break
    return rows[:limit]


async def _harvest_ukons(endpoint: str, limit: int, timeout_s: int) -> list[dict]:
    rows: list[dict] = []
    offset = 0
    per_page = 50
    timeout = aiohttp.ClientTimeout(total=timeout_s)
    async with aiohttp.ClientSession(timeout=timeout) as session:
        while len(rows) < limit:
            params = {"offset": offset, "limit": per_page}
            async with session.get(endpoint, params=params) as resp:
                if resp.status != 200:
                    break
                data = await resp.json(content_type=None)
            items = data.get("items", []) if isinstance(data, dict) else []
            if not items:
                break
            rows.extend([r for r in items if isinstance(r, dict)])
            offset += per_page
    return rows[:limit]


async def _harvest_undata(spec: SourceSpec, timeout_s: int) -> list[dict]:
    timeout = aiohttp.ClientTimeout(total=timeout_s)
    async with aiohttp.ClientSession(timeout=timeout) as session:
        async with session.get(spec.endpoint, headers={"Accept": "text/json"}) as resp:
            if resp.status != 200:
                return []
            payload = await resp.json(content_type=None)

    refs = payload.get("references", {}) if isinstance(payload, dict) else {}
    rows: list[dict] = []
    if isinstance(refs, dict):
        for value in refs.values():
            if isinstance(value, dict):
                rows.append(value)
            elif isinstance(value, list):
                rows.extend([v for v in value if isinstance(v, dict)])

    allow = set(spec.agency_allowlist)
    exclude = set(spec.exclude_agencies)
    filtered: list[dict] = []
    for row in rows:
        agency = str(row.get("agencyID", ""))
        if allow and agency not in allow:
            continue
        if agency in exclude:
            continue
        filtered.append(row)
    return filtered


async def _harvest_sdmx_dataflows(spec: SourceSpec, timeout_s: int) -> list[dict]:
    timeout = aiohttp.ClientTimeout(total=timeout_s)
    async with aiohttp.ClientSession(timeout=timeout) as session:
        async with session.get(spec.endpoint, headers={"Accept": "application/xml"}) as resp:
            if resp.status != 200:
                return []
            content = await resp.read()

    try:
        root = ET.fromstring(content)
    except ET.ParseError:
        return []

    rows: list[dict] = []
    for elem in root.iter():
        tag = elem.tag.split("}", 1)[1] if "}" in elem.tag else elem.tag
        if tag != "Dataflow":
            continue
        row = {
            "id": elem.attrib.get("id", ""),
            "agencyID": elem.attrib.get("agencyID", ""),
            "version": elem.attrib.get("version", ""),
            "name": "",
            "description": "",
        }
        for child in elem:
            ctag = child.tag.split("}", 1)[1] if "}" in child.tag else child.tag
            if ctag == "Name" and not row["name"]:
                row["name"] = (child.text or "").strip()
            elif ctag == "Description" and not row["description"]:
                row["description"] = (child.text or "").strip()

        agency = str(row.get("agencyID", ""))
        if spec.agency_prefix and not agency.startswith(spec.agency_prefix):
            continue
        rows.append(row)
    return rows


async def _harvest_who_indicators(endpoint: str, limit: int, timeout_s: int) -> list[dict]:
    rows: list[dict] = []
    next_url: str | None = endpoint
    timeout = aiohttp.ClientTimeout(total=timeout_s)

    async with aiohttp.ClientSession(timeout=timeout) as session:
        while next_url and len(rows) < limit:
            async with session.get(next_url, headers={"Accept": "application/json"}) as resp:
                if resp.status != 200:
                    break
                payload = await resp.json(content_type=None)
            batch = payload.get("value", []) if isinstance(payload, dict) else []
            if not isinstance(batch, list) or not batch:
                break
            rows.extend([row for row in batch if isinstance(row, dict)])
            next_url = payload.get("@odata.nextLink") if isinstance(payload, dict) else None
    return rows[:limit]


async def _harvest_uis_indicators(endpoint: str, limit: int, timeout_s: int) -> list[dict]:
    timeout = aiohttp.ClientTimeout(total=timeout_s)
    async with aiohttp.ClientSession(timeout=timeout) as session:
        async with session.get(endpoint, headers={"Accept": "application/json"}) as resp:
            if resp.status != 200:
                return []
            payload = await resp.json(content_type=None)

    if isinstance(payload, list):
        rows = [row for row in payload if isinstance(row, dict)]
    elif isinstance(payload, dict):
        if isinstance(payload.get("data"), list):
            rows = [row for row in payload["data"] if isinstance(row, dict)]
        elif isinstance(payload.get("items"), list):
            rows = [row for row in payload["items"] if isinstance(row, dict)]
        else:
            rows = [payload]
    else:
        rows = []
    return rows[:limit]


async def _harvest_wvs(endpoint: str, limit: int, timeout_s: int) -> list[dict]:
    del endpoint, timeout_s
    rows = [dict(item) for item in _load_wvs_indicator_catalog_from_local_file()]
    if not rows:
        rows = [dict(item) for item in _WVS_STATIC_INDICATORS]
    return rows[:limit]


async def _harvest_unpd_indicators(endpoint: str, limit: int, timeout_s: int) -> list[dict]:
    rows: list[dict] = []
    timeout = aiohttp.ClientTimeout(total=timeout_s)
    next_url: str | None = endpoint
    page = 1

    async with aiohttp.ClientSession(timeout=timeout) as session:
        while len(rows) < limit:
            if next_url:
                request_url = next_url
                params: dict[str, Any] | None = None
            else:
                request_url = endpoint
                params = {"page": page}
            async with session.get(
                request_url, params=params, headers={"Accept": "application/json"}
            ) as resp:
                if resp.status != 200:
                    break
                payload = await resp.json(content_type=None)

            if not isinstance(payload, dict):
                break
            batch = payload.get("data", [])
            if not isinstance(batch, list) or not batch:
                break
            rows.extend([row for row in batch if isinstance(row, dict)])

            next_url = (
                payload.get("next")
                or payload.get("nextPage")
                or payload.get("next_page")
                or payload.get("nextPageUrl")
            )
            if not next_url:
                page += 1
            if len(batch) == 0:
                break

    return rows[:limit]
