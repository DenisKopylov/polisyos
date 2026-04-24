#!/usr/bin/env python3
"""Generate WVS indicator registry YAML from Excel codebook + CSV data.

Reads the WVS Time Series Excel metadata file to get variable names, titles,
and wave availability, then samples the CSV to infer response types.

Usage:
    python scripts/generate_wvs_registry.py [--raw-dir data/raw/wvs]
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

# ---------------------------------------------------------------------------
# Response type inference
# ---------------------------------------------------------------------------


def _infer_response_type(values: list[float]) -> str:
    """Infer WVS response type from sampled valid values."""
    if not values:
        return "unknown"
    unique = sorted(set(values))
    min_v, max_v = unique[0], unique[-1]

    # Binary 0/1 (mentioned/not mentioned — child qualities etc.)
    if set(unique) <= {0.0, 1.0}:
        return "binary_mentioned"

    # Binary 1/2 (trust questions — "can be trusted" / "can't be too careful")
    if set(unique) <= {1.0, 2.0}:
        return "binary_12"

    # Likert 1-4 (importance, confidence, frequency)
    if min_v >= 1.0 and max_v <= 4.0 and all(v == int(v) for v in unique):
        return "likert_4"

    # Likert 1-5
    if min_v >= 1.0 and max_v <= 5.0 and all(v == int(v) for v in unique):
        return "likert_5"

    # Likert 1-10 (satisfaction, importance scales)
    if min_v >= 1.0 and max_v <= 10.0 and all(v == int(v) for v in unique):
        return "likert_10"

    # Continuous / composite index
    return "continuous"


# ---------------------------------------------------------------------------
# Canonical metric candidates (curated mapping)
# ---------------------------------------------------------------------------

_CANONICAL_CANDIDATES: dict[str, list[str]] = {
    # Social trust
    "A165": ["social_trust"],
    "A214": ["social_trust"],
    "G007_64": ["social_trust"],
    # Social capital / life satisfaction
    "A008": ["social_capital"],
    "A170": ["social_capital"],
    "A173": ["cultural_cluster"],
    "E023": ["social_capital"],
    "E025": ["social_capital"],
    "E286": ["social_capital"],
    # Important in life series
    "A001": ["social_capital"],
    "A002": ["social_capital"],
    "A003": ["social_capital"],
    "A004": ["social_capital"],
    "A005": ["social_capital"],
    "A006": ["social_capital"],
    # Health
    "A009": ["health_outcomes"],
    # Institutional confidence
    "E069_04": ["public_trust"],
    "E069_06": ["public_trust"],
    "E069_07": ["public_trust", "institutional_quality"],
    "E069_11": ["public_trust", "institutional_quality"],
    "E069_17": ["judicial_quality", "institutional_quality"],
    "E069_12": ["public_trust"],
    "E069_13": ["public_trust"],
    "E069_18": ["public_trust"],
    # Democracy
    "E110": ["democracy_quality"],
    "E117": ["democracy_quality"],
    "F108": ["democracy_quality"],
    # Corruption
    "E196": ["corruption_level"],
    # Gender equality
    "D059": ["gender_equality"],
    "D060": ["gender_equality"],
    "E233": ["gender_equality"],
    "Y022": ["gender_equality"],
    # Inequality
    "E035": ["inequality"],
    # Child qualities series
    "A029": ["social_capital"],
    "A030": ["social_capital"],
    "A032": ["social_capital"],
    "A034": ["social_capital"],
    "A035": ["social_capital"],
    "A038": ["social_capital"],
    "A039": ["social_capital"],
    "A040": ["social_capital"],
    "A041": ["social_capital"],
    "A042": ["social_capital"],
    # Political action
    "A062": ["social_capital"],
}

# Indicators where response 1 means "positive" and should use share aggregation
_BINARY_SHARE_INDICATORS = frozenset({"A165"})

# Metadata columns to skip
_METADATA_PREFIXES = (
    "S",
    "COUNTRY",
    "COW",
    "MODE",
    "CASEID",
    "N",
    "GWNO",
    "version",
    "doi",
    "survself",
    "tradrat",
    "TradAgg",
    "SurvSAgg",
)


def _is_indicator_column(col: str) -> bool:
    """Return True if column looks like a WVS indicator (not metadata)."""
    return not any(col.startswith(p) for p in _METADATA_PREFIXES)


def _top_level_code(col: str) -> str:
    """Extract top-level question code: E069_07 -> E069, A001 -> A001."""
    m = re.match(r"^([A-Z]+\d+)", col)
    return m.group(1) if m else col


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate WVS indicator registry")
    parser.add_argument("--raw-dir", type=Path, default=Path("data/raw/wvs"))
    parser.add_argument(
        "--output", type=Path, default=Path("data/dataset_catalog/wvs_indicator_registry.yaml")
    )
    parser.add_argument(
        "--sample-rows",
        type=int,
        default=8000,
        help="Number of CSV rows to sample for response type inference",
    )
    args = parser.parse_args()

    raw_dir = args.raw_dir
    csv_path = raw_dir / "WVS_Time_Series_1981-2022_csv_v5_0.csv"
    xlsx_path = (
        raw_dir / "F00003844-WVS_Time_Series_List_of_Variables_and_equivalences_1981_2022_v3_1.xlsx"
    )

    if not csv_path.exists():
        print(f"ERROR: CSV not found at {csv_path}", file=sys.stderr)
        sys.exit(1)

    # -----------------------------------------------------------------------
    # Step 1: Read Excel codebook for titles and wave availability
    # -----------------------------------------------------------------------
    codebook: dict[str, dict] = {}  # variable -> {title, waves}

    if xlsx_path.exists():
        try:
            from openpyxl import load_workbook

            wb = load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            # Header: None, Variable, Title, WVS7, WVS6, WVS5, WVS4, WVS3, WVS2, WVS1
            rows_iter = ws.iter_rows(values_only=True)
            header = next(rows_iter, ())
            # Find column indices
            var_idx = title_idx = None
            wave_indices: dict[int, int] = {}  # wave_number -> column_index
            for i, cell in enumerate(header):
                label = str(cell or "").strip()
                if label == "Variable":
                    var_idx = i
                elif label == "Title":
                    title_idx = i
                else:
                    m = re.match(r"WVS(\d+)", label)
                    if m:
                        wave_indices[int(m.group(1))] = i

            if var_idx is not None and title_idx is not None:
                for row in rows_iter:
                    variable = str(row[var_idx] or "").strip()
                    if not variable:
                        continue
                    title = str(row[title_idx] or "").strip()
                    waves = sorted(
                        wn
                        for wn, ci in wave_indices.items()
                        if row[ci] is not None and str(row[ci]).strip()
                    )
                    codebook[variable] = {"title": title, "waves": waves}
                print(f"Loaded {len(codebook)} indicators from Excel codebook")
            wb.close()
        except Exception as exc:
            print(f"WARNING: Could not read Excel codebook: {exc}", file=sys.stderr)

    # -----------------------------------------------------------------------
    # Step 2: Read CSV headers to discover all columns and sub-items
    # -----------------------------------------------------------------------
    with open(csv_path, encoding="utf-8") as f:
        reader = csv.reader(f)
        csv_headers = next(reader)

    indicator_cols = [h for h in csv_headers if _is_indicator_column(h)]
    print(f"Found {len(indicator_cols)} indicator columns in CSV")

    # Group sub-items by top-level code
    sub_items_map: dict[str, list[str]] = defaultdict(list)
    for col in indicator_cols:
        top = _top_level_code(col)
        if col != top:
            sub_items_map[top].append(col)

    # -----------------------------------------------------------------------
    # Step 3: Sample CSV to infer response types
    # -----------------------------------------------------------------------
    print(f"Sampling {args.sample_rows} rows to infer response types...")
    value_samples: dict[str, list[float]] = defaultdict(list)

    with open(csv_path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader):
            if i >= args.sample_rows:
                break
            for col in indicator_cols:
                raw = row.get(col, "")
                if not raw or raw.strip() == "":
                    continue
                try:
                    val = float(raw)
                except (ValueError, TypeError):
                    continue
                # Skip WVS missing codes
                if val < 0:
                    continue
                if len(value_samples[col]) < 2000:
                    value_samples[col].append(val)

    response_types: dict[str, str] = {}
    for col in indicator_cols:
        response_types[col] = _infer_response_type(value_samples.get(col, []))

    # -----------------------------------------------------------------------
    # Step 4: Build registry and write YAML
    # -----------------------------------------------------------------------
    # Determine which top-level indicators to include
    # An indicator is included if:
    # - It's in the codebook, OR
    # - It's a column in the CSV that is an indicator
    # Sub-items are grouped under their parent

    all_top_level = set()
    standalone_cols = []
    for col in indicator_cols:
        top = _top_level_code(col)
        if col == top:
            all_top_level.add(col)
            standalone_cols.append(col)
        else:
            all_top_level.add(top)

    # Also add codebook entries that might not be in CSV headers
    for var in codebook:
        if _is_indicator_column(var):
            all_top_level.add(var)

    # Sort indicators naturally
    def _sort_key(code: str) -> tuple:
        m = re.match(r"^([A-Z]+)(\d+)", code)
        if m:
            return (m.group(1), int(m.group(2)))
        return (code, 0)

    sorted_indicators = sorted(all_top_level, key=_sort_key)

    # Build YAML output
    lines: list[str] = []
    lines.append("# WVS Indicator Registry")
    lines.append("# Auto-generated from Excel codebook + CSV response type inference.")
    lines.append("# Curate canonical_candidates for high-priority indicators.")
    lines.append(f"# Total indicators: {len(sorted_indicators)}")
    lines.append("")
    lines.append("version: 1")
    lines.append("indicators:")

    stats = Counter()
    for code in sorted_indicators:
        cb = codebook.get(code, {})
        title = cb.get("title", "")
        waves = cb.get("waves", [])

        # Get response type — prefer the top-level column's type,
        # fall back to first sub-item's type
        rt = response_types.get(code, "")
        if not rt or rt == "unknown":
            subs = sub_items_map.get(code, [])
            for s in subs:
                srt = response_types.get(s, "")
                if srt and srt != "unknown":
                    rt = srt
                    break
        if not rt:
            rt = "unknown"

        stats[rt] += 1

        # Determine aggregation from response type
        if code in _BINARY_SHARE_INDICATORS or rt == "binary_12":
            aggregation = "weighted_share_response_1"
        elif rt == "binary_mentioned":
            aggregation = "weighted_share_mentioned"
        elif rt.startswith("likert") or rt == "continuous":
            aggregation = "weighted_mean"
        else:
            aggregation = "weighted_mean"

        subs = sub_items_map.get(code, [])
        candidates = _CANONICAL_CANDIDATES.get(code, [])

        lines.append(f"  {code}:")
        # Escape title for YAML
        safe_title = title.replace('"', '\\"') if title else code
        lines.append(f'    title: "{safe_title}"')
        lines.append(f"    response_type: {rt}")
        lines.append(f"    aggregation: {aggregation}")
        if waves:
            lines.append(f"    waves: [{', '.join(str(w) for w in waves)}]")
        if subs:
            lines.append(f"    sub_items: [{', '.join(sorted(subs, key=_sort_key))}]")
        if candidates:
            lines.append(f"    canonical_candidates: [{', '.join(candidates)}]")

    output_path = args.output
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")

    print(f"\nWritten {len(sorted_indicators)} indicators to {output_path}")
    print(f"Response type distribution: {dict(stats)}")
    print(
        f"Sub-item families: {len(sub_items_map)} (total sub-items: {sum(len(v) for v in sub_items_map.values())})"
    )
    print(
        f"Indicators with canonical candidates: {sum(1 for c in sorted_indicators if c in _CANONICAL_CANDIDATES)}"
    )


if __name__ == "__main__":
    main()
